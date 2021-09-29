from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
# Sheet Names
SheetNames = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
MasterCalendar = "Master.xlsx"
FacultyCalendar = "FacultyCalendar_Output.xlsx"
df = pd.read_excel(MasterCalendar, 'Key')
Initiatives = df['FixedInitiativeTitles'].dropna()
Initiativecode = df['FixedInitiativeCodes'].dropna()
Init = []
for i in range(len(Initiatives)):
    Init.append(Initiativecode[i]+'p')
    Init.append(Initiativecode[i]+'s')
#Running all the sheets of Master Calendar
for Sheet in range(12):
    MonthSheet = SheetNames[Sheet]
    wb1 = load_workbook(MasterCalendar)
    MasterExcel = wb1[MonthSheet]
    wb2 = load_workbook(FacultyCalendar)
    FacultyExcel = wb2[MonthSheet]
    FacultyExcel.delete_rows(5,FacultyExcel.max_row-4)
    #Loading all the rows and columns of particular Month sheet
    all_rows = list(MasterExcel.rows)
    all_Columns = list(MasterExcel.columns)
    FacultydList = []
    Leads1 = []
    Leads2 = []
    Leads3 = []
    Leads4 = []
    Leads5 = []
    for row in range(3,len(all_rows)):
        for column in range(7):
            if column==2:
                if all_rows[row][column].value:
                    #Getting the Primary Faculty Names
                    FacultydList.append((all_rows[row][column].value).upper())
                    Leads1.append((all_rows[row][column].value).upper())
                else:
                    Leads1.append((all_rows[row][column].value))
            if column==3:
                #Getting the Secondary Faculty Names
                if all_rows[row][column].value:
                    FacultydList.append((all_rows[row][column].value).upper())
                    Leads2.append((all_rows[row][column].value).upper())
                else:
                    Leads2.append(all_rows[row][column].value)
            if column==4:
                #Getting the Secondary Faculty Names
                if all_rows[row][column].value:
                    FacultydList.append((all_rows[row][column].value).upper())
                    Leads3.append((all_rows[row][column].value).upper())
                else:
                    Leads3.append(all_rows[row][column].value)
            if column==5:
                #Getting the Secondary Faculty Names
                if all_rows[row][column].value:
                    FacultydList.append((all_rows[row][column].value).upper())
                    Leads4.append((all_rows[row][column].value).upper())
                else:
                    Leads4.append(all_rows[row][column].value)
            if column==6:
                #Getting the Secondary Faculty Names
                if all_rows[row][column].value:
                    FacultydList.append((all_rows[row][column].value).upper())
                    Leads5.append((all_rows[row][column].value).upper())
                else:
                    Leads5.append(all_rows[row][column].value)
    FacultyList = []
    for val in FacultydList:
        if val != None :
            FacultyList.append(val)
    FacultyList = list(set(FacultyList))
    FacultySlots = []
    ExistingSlots = []
    for i in range(len(FacultyList)):
        FacultySlots.append(['']*124)
    for i in range(len(Leads1)):
        ExistingSlots.append(['']*124)

    for row in range(3,len(Leads1)+3):
        for column in range(7,130):
            if all_rows[row][column].value:
                ExistingSlots[row-3][column-7]=all_rows[row][column].value
    for i in range(len(Leads1)):
        for j in range(len(FacultyList)):
            if FacultyList[j]==Leads1[i]:
                for k in range(124):
                    if ExistingSlots[i][k]:
                        FacultySlots[j][k] = FacultySlots[j][k]+ str(ExistingSlots[i][k])+ 'p'
            if FacultyList[j]==Leads2[i]:
                for k in range(124):
                    if ExistingSlots[i][k]:
                        FacultySlots[j][k] = FacultySlots[j][k]+ str(ExistingSlots[i][k])+ 's'
            if FacultyList[j]==Leads3[i]:
                for k in range(124):
                    if ExistingSlots[i][k]:
                        FacultySlots[j][k] = FacultySlots[j][k]+ str(ExistingSlots[i][k])+ 's'
            if FacultyList[j]==Leads4[i]:
                for k in range(124):
                    if ExistingSlots[i][k]:
                        FacultySlots[j][k] = FacultySlots[j][k]+ str(ExistingSlots[i][k])+ 's'
            if FacultyList[j]==Leads5[i]:
                for k in range(124):
                    if ExistingSlots[i][k]:
                        FacultySlots[j][k] = FacultySlots[j][k]+ str(ExistingSlots[i][k])+ 's'
    for i in range(len(FacultyList)):
        FacultyExcel.cell(row = 5+i, column = 1).value = i+1
        FacultyExcel.cell(row = 5+i, column = 2).value = FacultyList[i]
        for j in range(124):
            FacultyExcel.cell(row = 5+i, column = 3+j).value = FacultySlots[i][j]
            correct = 0
            for ini in range(len(Initiatives)*2):
                if FacultySlots[i][j] == Init[ini]:
                    correct = 1
                    break
            if correct == 0:
                if FacultySlots[i][j]: 
                    FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color="FF0000", end_color="FF0000")
    wb2.save(FacultyCalendar)
                         
    
        
    
    
    
    

