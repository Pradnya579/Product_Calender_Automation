import pandas as pd
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    SheetName = "Sample_GENESIS"
    InputCalendar = "Automation_Sample Calender_v0.6.xlsx"
    Initiative = "GENESIS"
    Month = "July"
    Month = Month.upper()
    wb = load_workbook(InputCalendar)
    InputExcel = wb[SheetName]
    all_rows = list(InputExcel.rows)
    all_Columns = list(InputExcel.columns)
    df = pd.read_excel('Automation_Sample Calender_v0.6.xlsx', 'Key')
    Initiatives = df['FixedInitiativeTitles']
    Initiativecode = df['FixedInitiativeCodes']
    Date = []
    FacultyList = []
    Leads1 = []
    Leads2 = []
    Leads3 = []
    SessionSlot = []
    RowLength = len(all_Columns[1])-3
    for head in range(8):
        for cell in range(3, len(all_Columns[head+1])):
            if head+1 == 1:
                list.append(Date, all_Columns[head+1][cell].value)
            elif head+1 == 2 | head+1 == 3 | head+1 == 4:
                break
            elif head+1 == 5:
                list.append(FacultyList, all_Columns[head + 1][cell].value)
                list.append(Leads1,all_Columns[head+1][cell].value)
            elif head+1 == 6:
                list.append(FacultyList, all_Columns[head + 1][cell].value)
                list.append(Leads2, all_Columns[head + 1][cell].value)
            elif head+1 == 7:
                list.append(FacultyList, all_Columns[head+1][cell].value)
                list.append(Leads3, all_Columns[head + 1][cell].value)
            elif head+1 == 8:
                list.append(SessionSlot, all_Columns[head+1][cell].value)
    FacultyList = list(set(FacultyList))
    FacultyList = [i for i in FacultyList if i]
    M1Slots = [0]*(len(FacultyList))
    M2Slots = [0]*(len(FacultyList))
    A1Slots = [0]*(len(FacultyList))
    A2Slots = [0]*(len(FacultyList))
    for i in range(len(Leads1)):
        for j in range(len(FacultyList)):
            if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':
                if Leads1[i] == FacultyList[j]:
                    M1Slots[j] = M1Slots[j] + 1
                if Leads2[i] == FacultyList[j]:
                    M1Slots[j] = M1Slots[j] + 1
                if Leads3[i] == FacultyList[j]:
                    M1Slots[j] = M1Slots[j] + 1
            if SessionSlot[i] == 'M2' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':
                if Leads1[i] == FacultyList[j]:
                    M2Slots[j] = M2Slots[j] + 1
                if Leads2[i] == FacultyList[j]:
                    M2Slots[j] = M2Slots[j] + 1
                if Leads3[i] == FacultyList[j]:
                    M2Slots[j] = M2Slots[j] + 1
            if SessionSlot[i] == 'A1' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                if Leads1[i] == FacultyList[j]:
                    A1Slots[j] = A1Slots[j] + 1
                if Leads2[i] == FacultyList[j]:
                    A1Slots[j] = A1Slots[j] + 1
                if Leads3[i] == FacultyList[j]:
                    A1Slots[j] = A1Slots[j] + 1
            if SessionSlot[i] == 'A2' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                if Leads1[i] == FacultyList[j]:
                    A2Slots[j] = A2Slots[j] + 1
                if Leads2[i] == FacultyList[j]:
                    A2Slots[j] = A2Slots[j] + 1
                if Leads3[i] == FacultyList[j]:
                    A2Slots[j] = A2Slots[j] + 1
    Months = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER",
              "NOVEMBER", "DECEMBER"];
    for i in range(12):
        if Month == Months[i]:
            index = i
    OutputExcelSheet = 'FacultyLoadSheet_Output.xlsx'
    OutputSheet = pd.ExcelFile('FacultyLoadSheet_Output.xlsx')
    OutputSheetNames = OutputSheet.sheet_names
    wb = load_workbook(OutputExcelSheet)
    OutputExcel = wb[OutputSheetNames[index]]
    OutputRows = list(OutputExcel.rows)
    ExFacultyList = []
    for head in range(4,24):
        list.append(ExFacultyList,OutputRows[head][1].value)
    ExFacultyList = list(set(ExFacultyList))
    ExFacultyList = [i for i in ExFacultyList if i]
    CheckList = [0]*len(FacultyList)
    CheckList1 = [0]*len(ExFacultyList)
    for i in range(len(FacultyList)):
        for j in range(len(ExFacultyList)):
            if FacultyList[i] == ExFacultyList[j]:
                CheckList[i] = 1;
                CheckList1[j] = 1;
                for k in range(len(Initiatives)):
                    if Initiatives[k] == Initiative:
                        OutputExcel.cell(row=5+j, column=3+(k*4)).value = M1Slots[i]
                        OutputExcel.cell(row=5+j, column=4+(k*4)).value = M2Slots[i]
                        OutputExcel.cell(row=5+j, column=5+(k*4)).value = A1Slots[i]
                        OutputExcel.cell(row=5+j, column=6+(k*4)).value = A2Slots[i]
    ExFacLen = len(ExFacultyList)
    p=0
    for i in range(len(FacultyList)):
        if CheckList[i] == 0:
            OutputExcel.cell(row=5+ExFacLen+p, column=2).value = FacultyList[i]
            for k in range(len(Initiatives)):
                    if Initiatives[k] == Initiative:
                        OutputExcel.cell(row=5+ExFacLen+p, column=3+(k*4)).value = M1Slots[i]
                        OutputExcel.cell(row=5+ExFacLen+p, column=4+(k*4)).value = M2Slots[i]
                        OutputExcel.cell(row=5+ExFacLen+p, column=5+(k*4)).value = A1Slots[i]
                        OutputExcel.cell(row=5+ExFacLen+p, column=6+(k*4)).value = A2Slots[i]
            p=p+1
    wb.save(OutputExcelSheet)
    OutputExcelSheet = 'FacultyLoadSheet_Output.xlsx'
    OutputSheet = pd.ExcelFile('FacultyLoadSheet_Output.xlsx')
    OutputSheetNames = OutputSheet.sheet_names
    wb = load_workbook(OutputExcelSheet)
    OutputExcel = wb[OutputSheetNames[index]]
    OutputRows = list(OutputExcel.rows)
    UpFacultyList = []
    for head in range(4,24):
        list.append(UpFacultyList,OutputRows[head][1].value)
    UpFacultyList = list(set(UpFacultyList))
    UpFacultyList = [i for i in UpFacultyList if i]
    AvailableSlots = []
    AvailableSlots.append([])
    AvailableSlots.append([])
    AvailableSlots.append([])
    AvailableSlots.append([])
    OccupiedSlots = []
    OccupiedSlots.append([])
    OccupiedSlots.append([])
    OccupiedSlots.append([])
    OccupiedSlots.append([])
    for i in range(4,len(UpFacultyList)+4):
        sum1 = 0
        sum2 = 0
        sum3 = 0
        sum4 = 0
        for cell in range(len(Initiatives)):
            if OutputRows[i][cell*4+3].value:
                sum1 = sum1+ OutputRows[i][cell*4+2].value
                sum2 = sum2+ OutputRows[i][cell*4+3].value
                sum3 = sum3+ OutputRows[i][cell*4+4].value
                sum4 = sum4+ OutputRows[i][cell*4+5].value
        list.append(OccupiedSlots[0],sum1)
        list.append(OccupiedSlots[1],sum2)
        list.append(OccupiedSlots[2],sum3)
        list.append(OccupiedSlots[3],sum4)
    for i in range(len(UpFacultyList)):
        AvailableSlots[0].append(18-OccupiedSlots[0][i])
        AvailableSlots[1].append(18-OccupiedSlots[1][i])
        AvailableSlots[2].append(18-OccupiedSlots[2][i])
        AvailableSlots[3].append(18-OccupiedSlots[3][i])
    for i in range(len(UpFacultyList)):
        OutputExcel.cell(row=5+i, column=31).value = 18
        OutputExcel.cell(row=5+i, column=32).value = 18
        OutputExcel.cell(row=5+i, column=33).value = 18
        OutputExcel.cell(row=5+i, column=34).value = 18
        OutputExcel.cell(row=5+i, column=35).value = AvailableSlots[0][i]
        OutputExcel.cell(row=5+i, column=36).value = AvailableSlots[1][i]
        OutputExcel.cell(row=5+i, column=37).value = AvailableSlots[2][i]
        OutputExcel.cell(row=5+i, column=38).value = AvailableSlots[3][i]
    wb.save(OutputExcelSheet)
    plt.bar(UpFacultyList, OccupiedSlots[0])
    plt.xlabel("Faculty")
    plt.ylabel("Morning Occupied Slots")
    temp = OutputSheetNames[index] + " Month Faculty wise Load (M)"
    plt.title(temp)
    plt.show()
    plt.bar(UpFacultyList, OccupiedSlots[1])
    plt.xlabel("Faculty")
    plt.ylabel("Afternoon Occupied Slots")
    temp = OutputSheetNames[index] + " Month Faculty wise Load (A)"
    plt.title(temp)
    plt.show()
    
                
                
    
        
                    
                
